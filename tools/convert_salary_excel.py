#!/usr/bin/env python3
"""
Convert salary data from Excel to JSON format.

This script converts an Excel file containing company salary data
into the JSON format expected by salary_lookup.py.

Prerequisites:
    pip install openpyxl

Usage:
    python tools/convert_salary_excel.py <path-to-excel-file>
    python tools/convert_salary_excel.py <path-to-excel-file> --source "My Union Stats 2025"
    python tools/convert_salary_excel.py <path-to-excel-file> --baseline 100 --baseline-desc "Index 100 = median salary"

The output file (salary_data.json) will be written to the repository root.

Expected Excel format:
    - A header row with column names
    - A "Company" or "Firma" column (required)
    - An optional "City" or "By" column
    - Any number of numeric data columns (salary index, count, etc.)

The script auto-detects the header row and column layout. For Excel files
with paired count/index columns per category, it groups them automatically.
"""

import json
import sys
import argparse
import re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None


# Column name patterns for auto-detection
COMPANY_PATTERNS = {"firma", "company", "virksomhed", "employer", "arbejdsgiver"}
CITY_PATTERNS = {"by", "city", "kommune", "location", "lokation", "sted"}
COUNT_PATTERNS = {"antal", "count", "number", "n", "employees", "medarbejdere"}
INDEX_PATTERNS = {"indeks", "index", "idx", "salary", "løn", "median", "average", "gennemsnit"}
# "Compound" tokens: pattern words allowed to match as a substring of a larger
# header token, for languages that glue words together (e.g. Danish "lønindeks"
# -> løn + indeks). Languages that write headers as separate words need none.
# Ships populated for this repo's Danish demonstration data; a fork targeting
# another locale edits this constant.
COMPOUND_PATTERNS = {"antal", "indeks", "løn", "gennemsnit", "medarbejdere"}
# Identifier columns (employee id, Danish "personnummer", etc.) are never salary
# data. They are dropped at classification so they are not mistaken for a salary
# category. Matched as whole tokens only, like other pattern sets.
ID_PATTERNS = {"id", "personnummer"}


def header_matches(header, patterns):
    """Return True when a header contains a meaningful pattern match.

    Patterns match whole tokens; any pattern also listed in
    ``COMPOUND_PATTERNS`` may additionally match as a substring, to handle
    languages that form compound words.
    """
    h = header.lower().strip()
    tokens = set(re.findall(r"[a-zæøåöäü0-9]+", h))

    for p in patterns:
        if p in tokens:
            return True
        if p in COMPOUND_PATTERNS and p in h:
            return True
    return False


def strip_type_patterns(header, patterns):
    """Remove count/index words from a header to derive a category name."""
    name = header.lower()
    for p in patterns:
        name = re.sub(rf"(?<![a-zæøåöäü0-9]){re.escape(p)}(?![a-zæøåöäü0-9])", "", name)
    return name.strip(" _-")


def detect_column_type(header):
    """Detect whether a column header refers to count or index data."""
    if header_matches(header, COUNT_PATTERNS):
        return "count"
    if header_matches(header, INDEX_PATTERNS):
        return "index"
    return None


def parse_sheet(ws, sheet_label=None):
    """Parse a single worksheet into a list of company entries and detected categories."""
    # Find header row
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=False), start=1):
        for cell in row:
            if cell.value and header_matches(str(cell.value), COMPANY_PATTERNS):
                header_row = row_idx
                break
        if header_row:
            break

    if header_row is None:
        print(f"Warning: Could not find header row in sheet '{ws.title}'. Skipping.", file=sys.stderr)
        return []

    # Read headers
    headers = []
    for cell in ws[header_row]:
        headers.append(str(cell.value).strip() if cell.value else "")

    # Find company and city columns
    company_col = None
    city_col = None
    for i, h in enumerate(headers):
        if header_matches(h, COMPANY_PATTERNS):
            company_col = i
        elif header_matches(h, CITY_PATTERNS):
            city_col = i

    if company_col is None:
        print(f"Warning: Could not find company column in sheet '{ws.title}'.", file=sys.stderr)
        return []

    # Identify data columns (everything that's not company/city or an identifier)
    data_cols = []
    for i, h in enumerate(headers):
        if i == company_col or i == city_col or not h:
            continue
        if header_matches(h, ID_PATTERNS):
            continue
        data_cols.append((i, h))

    # Group data columns by detected type and derive category names
    count_cols = []
    index_cols = []
    untyped_cols = []

    for col_idx, col_header in data_cols:
        col_type = detect_column_type(col_header)
        if col_type == "count":
            cat_name = strip_type_patterns(col_header, COUNT_PATTERNS)
            count_cols.append((col_idx, col_header, cat_name))
        elif col_type == "index":
            cat_name = strip_type_patterns(col_header, INDEX_PATTERNS)
            index_cols.append((col_idx, col_header, cat_name))
        else:
            untyped_cols.append((col_idx, col_header))

    # Pair count/index columns by matching category name
    categories = []
    used_counts = set()
    used_indexes = set()

    for ci, (c_idx, c_header, c_cat) in enumerate(count_cols):
        for ii, (i_idx, i_header, i_cat) in enumerate(index_cols):
            if ii in used_indexes:
                continue
            if c_cat and i_cat and c_cat == i_cat:
                cat_name = c_cat.replace(" ", "_").replace("-", "_")
                categories.append({
                    "name": cat_name,
                    "count_col": c_idx,
                    "index_col": i_idx,
                })
                used_counts.add(ci)
                used_indexes.add(ii)
                break

    # Remaining unmatched count columns become standalone. They are still count
    # data, so tag them as such — otherwise a lone headcount would be emitted as
    # a salary index and rendered with a meaningless "vs baseline" percentage.
    for ci, (c_idx, c_header, _) in enumerate(count_cols):
        if ci not in used_counts:
            categories.append(
                {"name": c_header.lower().replace(" ", "_"), "value_col": c_idx, "field": "count"}
            )

    # Remaining unmatched index columns become standalone (use original header)
    for ii, (i_idx, i_header, _) in enumerate(index_cols):
        if ii not in used_indexes:
            categories.append({"name": i_header.lower().replace(" ", "_"), "value_col": i_idx})

    # Untyped columns become standalone
    for col_idx, col_header in untyped_cols:
        categories.append({"name": col_header.lower().replace(" ", "_"), "value_col": col_idx})

    # Parse data rows
    companies = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row[company_col]:
            continue

        company_name = str(row[company_col]).strip()
        city_name = str(row[city_col]).strip() if city_col is not None and row[city_col] else ""

        entry = {
            "company": company_name,
            "city": city_name,
            "categories": {},
        }

        for cat in categories:
            cat_name = cat["name"]
            if "count_col" in cat and "index_col" in cat:
                count_val = None
                index_val = None
                if cat["count_col"] < len(row) and row[cat["count_col"]] is not None:
                    try:
                        count_val = int(row[cat["count_col"]])
                    except (ValueError, TypeError):
                        pass
                if cat["index_col"] < len(row) and row[cat["index_col"]] is not None:
                    try:
                        index_val = float(row[cat["index_col"]])
                    except (ValueError, TypeError):
                        pass
                # A count/index pair that is entirely empty for this row carries
                # no salary information, so skip it rather than emit nulls.
                if count_val is None and index_val is None:
                    continue
                entry["categories"][cat_name] = {"count": count_val, "index": index_val}
            elif "value_col" in cat:
                if cat["value_col"] < len(row) and row[cat["value_col"]] is not None:
                    val = row[cat["value_col"]]
                    try:
                        val = float(val)
                    except (ValueError, TypeError):
                        # Non-numeric standalone value (e.g. a free-text "Notes"
                        # column) is not salary data; skip it for this row.
                        continue
                    field = cat.get("field", "index")
                    entry["categories"][cat_name] = {field: int(val) if field == "count" else val}

        companies.append(entry)

    return companies


def main():
    parser = argparse.ArgumentParser(
        description="Convert salary Excel data to JSON"
    )
    parser.add_argument("excel_file", help="Path to the Excel file with salary data")
    parser.add_argument(
        "--output", default=None,
        help="Output JSON file path (default: salary_data.json in repo root)",
    )
    parser.add_argument(
        "--source", default=None,
        help="Name of the data source (e.g., 'Union Statistics 2025')",
    )
    parser.add_argument(
        "--baseline", type=float, default=100,
        help="Baseline value for index comparison (default: 100)",
    )
    parser.add_argument(
        "--baseline-desc", default=None,
        help="Description of what the baseline means (e.g., 'Index 100 = median salary')",
    )
    args = parser.parse_args()

    excel_path = Path(args.excel_file)
    if not excel_path.exists():
        print(f"Error: File not found: {excel_path}", file=sys.stderr)
        sys.exit(1)

    if openpyxl is None:
        print("Error: openpyxl is required. Install it with: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    output_path = Path(args.output) if args.output else Path(__file__).parent.parent / "salary_data.json"

    print(f"Reading: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    all_companies = []
    for sheet_name in wb.sheetnames:
        print(f"  Parsing sheet: {sheet_name}")
        ws = wb[sheet_name]
        companies = parse_sheet(ws, sheet_label=sheet_name)
        all_companies.extend(companies)

    wb.close()

    if not all_companies:
        print("Error: No data could be parsed from the Excel file.", file=sys.stderr)
        print("Make sure the Excel file has a header row with a 'Company'/'Firma' column.", file=sys.stderr)
        sys.exit(1)

    # Build output
    output = {
        "metadata": {
            "source": args.source or excel_path.stem,
            "index_baseline": args.baseline,
            "index_label": "Index",
            "baseline_description": args.baseline_desc or f"Index {args.baseline} = baseline",
        },
        "companies": all_companies,
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\nDone! Wrote {len(all_companies)} company entries to {output_path}")


if __name__ == "__main__":
    main()
